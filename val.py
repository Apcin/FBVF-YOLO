# Ultralytics YOLOv5 🚀, AGPL-3.0 license
"""
Validate a trained YOLOv5 detection model on a detection dataset.

Usage:
    $ python val.py --weights yolov5s.pt --data coco128.yaml --img 640

Usage - formats:
    $ python val.py --weights yolov5s.pt                 # PyTorch
                              yolov5s.torchscript        # TorchScript
                              yolov5s.onnx               # ONNX Runtime or OpenCV DNN with --dnn
                              yolov5s_openvino_model     # OpenVINO
                              yolov5s.engine             # TensorRT
                              yolov5s.mlpackage          # CoreML (macOS-only)
                              yolov5s_saved_model        # TensorFlow SavedModel
                              yolov5s.pb                 # TensorFlow GraphDef
                              yolov5s.tflite             # TensorFlow Lite
                              yolov5s_edgetpu.tflite     # TensorFlow Edge TPU
                              yolov5s_paddle_model       # PaddlePaddle
"""

import argparse
import json
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
import torch
from tqdm import tqdm

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

from models.common import DetectMultiBackend
from utils.callbacks import Callbacks
from utils.dataloaders import create_dataloader
from utils.general import (
    LOGGER,
    TQDM_BAR_FORMAT,
    Profile,
    check_dataset,
    check_img_size,
    check_requirements,
    check_yaml,
    coco80_to_coco91_class,
    colorstr,
    increment_path,
    non_max_suppression,
    print_args,
    scale_boxes,
    xywh2xyxy,
    xyxy2xywh,
)
from utils.metrics import ConfusionMatrix, ap_per_class, box_iou
from utils.plots import output_to_target, plot_images, plot_val_study
from utils.torch_utils import select_device, smart_inference_mode


def save_one_txt(predn, save_conf, shape, file):
    """
    Saves one detection result to a txt file in normalized xywh format, optionally including confidence.

    Args:
        predn (torch.Tensor): Predicted bounding boxes and associated confidence scores and classes in xyxy format, tensor
            of shape (N, 6) where N is the number of detections.
        save_conf (bool): If True, saves the confidence scores along with the bounding box coordinates.
        shape (tuple): Shape of the original image as (height, width).
        file (str | Path): File path where the result will be saved.

    Returns:
        None

    Notes:
        The xyxy bounding box format represents the coordinates (xmin, ymin, xmax, ymax).
        The xywh format represents the coordinates (center_x, center_y, width, height) and is normalized by the width and
        height of the image.

    Example:
        ```python
        predn = torch.tensor([[10, 20, 30, 40, 0.9, 1]])  # example prediction
        save_one_txt(predn, save_conf=True, shape=(640, 480), file="output.txt")
        ```
    """
    gn = torch.tensor(shape)[[1, 0, 1, 0]]  # normalization gain whwh
    for *xyxy, conf, cls in predn.tolist():
        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
        with open(file, "a") as f:
            f.write(("%g " * len(line)).rstrip() % line + "\n")


def save_one_json(predn, jdict, path, class_map):
    """
    Saves a single JSON detection result, including image ID, category ID, bounding box, and confidence score.

    Args:
        predn (torch.Tensor): Predicted detections in xyxy format with shape (n, 6) where n is the number of detections.
                              The tensor should contain [x_min, y_min, x_max, y_max, confidence, class_id] for each detection.
        jdict (list[dict]): List to collect JSON formatted detection results.
        path (pathlib.Path): Path object of the image file, used to extract image_id.
        class_map (dict[int, int]): Mapping from model class indices to dataset-specific category IDs.

    Returns:
        None: Appends detection results as dictionaries to `jdict` list in-place.

    Example:
        ```python
        predn = torch.tensor([[100, 50, 200, 150, 0.9, 0], [50, 30, 100, 80, 0.8, 1]])
        jdict = []
        path = Path("42.jpg")
        class_map = {0: 18, 1: 19}
        save_one_json(predn, jdict, path, class_map)
        ```
        This will append to `jdict`:
        ```
        [
            {'image_id': 42, 'category_id': 18, 'bbox': [125.0, 75.0, 100.0, 100.0], 'score': 0.9},
            {'image_id': 42, 'category_id': 19, 'bbox': [75.0, 55.0, 50.0, 50.0], 'score': 0.8}
        ]
        ```

    Notes:
        The `bbox` values are formatted as [x, y, width, height], where x and y represent the top-left corner of the box.
    """
    image_id = int(path.stem) if path.stem.isnumeric() else path.stem
    box = xyxy2xywh(predn[:, :4])  # xywh
    box[:, :2] -= box[:, 2:] / 2  # xy center to top-left corner
    for p, b in zip(predn.tolist(), box.tolist()):
        jdict.append(
            {
                "image_id": image_id,
                "category_id": class_map[int(p[5])],
                "bbox": [round(x, 3) for x in b],
                "score": round(p[4], 5),
            }
        )


def race_match_stats(detections, labels, nc, vehicle_cls=24, vehicle_iou=0.35, default_iou=0.50):
    """Count competition TP/FP/FN and error details with class-exact matching.

    Detections are [x1, y1, x2, y2, conf, class]; labels are [class, x1, y1, x2, y2].
    Matching is confidence-ordered, class-exact, one-to-one, with IoU 0.35 for vehicle class and 0.50 otherwise.
    """
    stats = {
        "tp": np.zeros(nc, dtype=np.int64),
        "fp": np.zeros(nc, dtype=np.int64),
        "fn": np.zeros(nc, dtype=np.int64),
        "gt": np.zeros(nc, dtype=np.int64),
        "fp_wrong_class": np.zeros(nc, dtype=np.int64),
        "fp_duplicate": np.zeros(nc, dtype=np.int64),
        "fp_background_or_iou": np.zeros(nc, dtype=np.int64),
        "confusion": np.zeros((nc, nc), dtype=np.int64),  # rows=true class, cols=pred class
    }
    nl, nd = labels.shape[0], detections.shape[0]

    for cls in labels[:, 0].int().tolist() if nl else []:
        stats["gt"][cls] += 1

    if nl == 0:
        for cls in detections[:, 5].int().tolist() if nd else []:
            stats["fp"][cls] += 1
            stats["fp_background_or_iou"][cls] += 1
        return stats
    if nd == 0:
        for cls in labels[:, 0].int().tolist():
            stats["fn"][cls] += 1
        return stats

    order = detections[:, 4].argsort(descending=True)
    ious = box_iou(labels[:, 1:], detections[:, :4])
    matched = torch.zeros(nl, dtype=torch.bool, device=labels.device)
    thresholds = torch.where(
        labels[:, 0] == vehicle_cls,
        torch.tensor(vehicle_iou, device=labels.device),
        torch.tensor(default_iou, device=labels.device),
    )

    for di in order.tolist():
        det_cls = int(detections[di, 5].item())
        same_cls = labels[:, 0] == det_cls
        candidates = torch.where(same_cls & ~matched & (ious[:, di] >= thresholds))[0]
        if candidates.numel():
            best = candidates[ious[candidates, di].argmax()]
            matched[best] = True
            stats["tp"][det_cls] += 1
            continue

        stats["fp"][det_cls] += 1
        wrong_cls = torch.where((labels[:, 0] != det_cls) & ~matched & (ious[:, di] >= thresholds))[0]
        if wrong_cls.numel():
            best = wrong_cls[ious[wrong_cls, di].argmax()]
            true_cls = int(labels[best, 0].item())
            stats["fp_wrong_class"][det_cls] += 1
            stats["confusion"][true_cls, det_cls] += 1
        elif torch.any(same_cls & matched & (ious[:, di] >= thresholds)):
            stats["fp_duplicate"][det_cls] += 1
        else:
            stats["fp_background_or_iou"][det_cls] += 1

    for cls in labels[~matched, 0].int().tolist():
        stats["fn"][cls] += 1
    return stats


def add_race_stats(total, update):
    """Accumulate per-image race statistics in-place."""
    for k, v in update.items():
        total[k] += v


def race_summary(tp, fp, fn):
    gt = tp + fn
    recall = tp / gt if gt else 0.0
    fdr = fp / (tp + fp) if tp + fp else 0.0
    return gt, recall, fdr


def process_batch(detections, labels, iouv):
    """
    Return a correct prediction matrix given detections and labels at various IoU thresholds.

    Args:
        detections (np.ndarray): Array of shape (N, 6) where each row corresponds to a detection with format
            [x1, y1, x2, y2, conf, class].
        labels (np.ndarray): Array of shape (M, 5) where each row corresponds to a ground truth label with format
            [class, x1, y1, x2, y2].
        iouv (np.ndarray): Array of IoU thresholds to evaluate at.

    Returns:
        correct (np.ndarray): A binary array of shape (N, len(iouv)) indicating whether each detection is a true positive
            for each IoU threshold. There are 10 IoU levels used in the evaluation.

    Example:
        ```python
        detections = np.array([[50, 50, 200, 200, 0.9, 1], [30, 30, 150, 150, 0.7, 0]])
        labels = np.array([[1, 50, 50, 200, 200]])
        iouv = np.linspace(0.5, 0.95, 10)
        correct = process_batch(detections, labels, iouv)
        ```

    Notes:
        - This function is used as part of the evaluation pipeline for object detection models.
        - IoU (Intersection over Union) is a common evaluation metric for object detection performance.
    """
    correct = np.zeros((detections.shape[0], iouv.shape[0])).astype(bool)
    iou = box_iou(labels[:, 1:], detections[:, :4])
    correct_class = labels[:, 0:1] == detections[:, 5]
    for i in range(len(iouv)):
        x = torch.where((iou >= iouv[i]) & correct_class)  # IoU > threshold and classes match
        if x[0].shape[0]:
            matches = torch.cat((torch.stack(x, 1), iou[x[0], x[1]][:, None]), 1).cpu().numpy()  # [label, detect, iou]
            if x[0].shape[0] > 1:
                matches = matches[matches[:, 2].argsort()[::-1]]
                matches = matches[np.unique(matches[:, 1], return_index=True)[1]]
                # matches = matches[matches[:, 2].argsort()[::-1]]
                matches = matches[np.unique(matches[:, 0], return_index=True)[1]]
            correct[matches[:, 1].astype(int), i] = True
    return torch.tensor(correct, dtype=torch.bool, device=iouv.device)


@smart_inference_mode()
def run(
    data,
    weights=None,  # model.pt path(s)
    batch_size=32,  # batch size
    imgsz=640,  # inference size (pixels)
    conf_thres=0.001,  # confidence threshold
    iou_thres=0.6,  # NMS IoU threshold
    max_det=300,  # maximum detections per image
    task="val",  # train, val, test, speed or study
    device="",  # cuda device, i.e. 0 or 0,1,2,3 or cpu
    workers=8,  # max dataloader workers (per RANK in DDP mode)
    single_cls=False,  # treat as single-class dataset
    augment=False,  # augmented inference
    verbose=False,  # verbose output
    save_txt=False,  # save results to *.txt
    save_hybrid=False,  # save label+prediction hybrid results to *.txt
    save_conf=False,  # save confidences in --save-txt labels
    save_json=False,  # save a COCO-JSON results file
    project=ROOT / "runs/val",  # save to project/name
    name="exp",  # save to project/name
    exist_ok=False,  # existing project/name ok, do not increment
    half=True,  # use FP16 half-precision inference
    dnn=False,  # use OpenCV DNN for ONNX inference
    model=None,
    dataloader=None,
    save_dir=Path(""),
    plots=True,
    callbacks=Callbacks(),
    compute_loss=None,
):
    """
    Evaluates a YOLOv5 model on a dataset and logs performance metrics.

    Args:
        data (str | dict): Path to a dataset YAML file or a dataset dictionary.
        weights (str | list[str], optional): Path to the model weights file(s). Supports various formats including PyTorch,
            TorchScript, ONNX, OpenVINO, TensorRT, CoreML, TensorFlow SavedModel, TensorFlow GraphDef, TensorFlow Lite,
            TensorFlow Edge TPU, and PaddlePaddle.
        batch_size (int, optional): Batch size for inference. Default is 32.
        imgsz (int, optional): Input image size (pixels). Default is 640.
        conf_thres (float, optional): Confidence threshold for object detection. Default is 0.001.
        iou_thres (float, optional): IoU threshold for Non-Maximum Suppression (NMS). Default is 0.6.
        max_det (int, optional): Maximum number of detections per image. Default is 300.
        task (str, optional): Task type - 'train', 'val', 'test', 'speed', or 'study'. Default is 'val'.
        device (str, optional): Device to use for computation, e.g., '0' or '0,1,2,3' for CUDA or 'cpu' for CPU. Default is ''.
        workers (int, optional): Number of dataloader workers. Default is 8.
        single_cls (bool, optional): Treat dataset as a single class. Default is False.
        augment (bool, optional): Enable augmented inference. Default is False.
        verbose (bool, optional): Enable verbose output. Default is False.
        save_txt (bool, optional): Save results to *.txt files. Default is False.
        save_hybrid (bool, optional): Save label and prediction hybrid results to *.txt files. Default is False.
        save_conf (bool, optional): Save confidences in --save-txt labels. Default is False.
        save_json (bool, optional): Save a COCO-JSON results file. Default is False.
        project (str | Path, optional): Directory to save results. Default is ROOT/'runs/val'.
        name (str, optional): Name of the run. Default is 'exp'.
        exist_ok (bool, optional): Overwrite existing project/name without incrementing. Default is False.
        half (bool, optional): Use FP16 half-precision inference. Default is True.
        dnn (bool, optional): Use OpenCV DNN for ONNX inference. Default is False.
        model (torch.nn.Module, optional): Model object for training. Default is None.
        dataloader (torch.utils.data.DataLoader, optional): Dataloader object. Default is None.
        save_dir (Path, optional): Directory to save results. Default is Path('').
        plots (bool, optional): Plot validation images and metrics. Default is True.
        callbacks (utils.callbacks.Callbacks, optional): Callbacks for logging and monitoring. Default is Callbacks().
        compute_loss (function, optional): Loss function for training. Default is None.

    Returns:
        dict: Contains performance metrics including precision, recall, mAP50, and mAP50-95.
    """
    # Initialize/load model and set device
    training = model is not None
    if training:  # called by train.py
        device, pt, jit, engine = next(model.parameters()).device, True, False, False  # get model device, PyTorch model
        half &= device.type != "cpu"  # half precision only supported on CUDA
        model.half() if half else model.float()
    else:  # called directly
        device = select_device(device, batch_size=batch_size)

        # Directories
        save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
        (save_dir / "labels" if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

        # Load model
        model = DetectMultiBackend(weights, device=device, dnn=dnn, data=data, fp16=half)
        stride, pt, jit, engine = model.stride, model.pt, model.jit, model.engine
        imgsz = check_img_size(imgsz, s=stride)  # check image size
        half = model.fp16  # FP16 supported on limited backends with CUDA
        if engine:
            batch_size = model.batch_size
        else:
            device = model.device
            if not (pt or jit):
                batch_size = 1  # export.py models default to batch-size 1
                LOGGER.info(f"Forcing --batch-size 1 square inference (1,3,{imgsz},{imgsz}) for non-PyTorch models")

        # Data
        data = check_dataset(data)  # check

    # Configure
    model.eval()
    cuda = device.type != "cpu"
    is_coco = isinstance(data.get("val"), str) and data["val"].endswith(f"coco{os.sep}val2017.txt")  # COCO dataset
    nc = 1 if single_cls else int(data["nc"])  # number of classes
    iouv = torch.linspace(0.5, 0.95, 10, device=device)  # iou vector for mAP@0.5:0.95
    niou = iouv.numel()

    # Dataloader
    if not training:
        if pt and not single_cls:  # check --weights are trained on --data
            ncm = model.model.nc
            assert ncm == nc, (
                f"{weights} ({ncm} classes) trained on different --data than what you passed ({nc} "
                f"classes). Pass correct combination of --weights and --data that are trained together."
            )
        model.warmup(imgsz=(1 if pt else batch_size, 3, imgsz, imgsz))  # warmup
        pad, rect = (0.0, False) if task == "speed" else (0.5, pt)  # square inference for benchmarks
        task = task if task in ("train", "val", "test") else "val"  # path to train/val/test images
        dataloader = create_dataloader(
            data[task],
            imgsz,
            batch_size,
            stride,
            single_cls,
            pad=pad,
            rect=rect,
            workers=workers,
            prefix=colorstr(f"{task}: "),
        )[0]

    seen = 0
    confusion_matrix = ConfusionMatrix(nc=nc)
    names = model.names if hasattr(model, "names") else model.module.names  # get class names
    if isinstance(names, (list, tuple)):  # old format
        names = dict(enumerate(names))
    class_map = coco80_to_coco91_class() if is_coco else list(range(1000))
    s = ("%22s" + "%11s" * 6) % ("Class", "Images", "Instances", "P", "R", "mAP50", "mAP50-95")
    tp, fp, p, r, f1, mp, mr, map50, ap50, map = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
    dt = Profile(device=device), Profile(device=device), Profile(device=device)  # profiling times
    loss = torch.zeros(3, device=device)
    jdict, stats, ap, ap_class = [], [], [], []
    race_stats = {
        "tp": np.zeros(nc, dtype=np.int64),
        "fp": np.zeros(nc, dtype=np.int64),
        "fn": np.zeros(nc, dtype=np.int64),
        "gt": np.zeros(nc, dtype=np.int64),
        "fp_wrong_class": np.zeros(nc, dtype=np.int64),
        "fp_duplicate": np.zeros(nc, dtype=np.int64),
        "fp_background_or_iou": np.zeros(nc, dtype=np.int64),
        "confusion": np.zeros((nc, nc), dtype=np.int64),
    }
    callbacks.run("on_val_start")
    pbar = tqdm(dataloader, desc=s, bar_format=TQDM_BAR_FORMAT)  # progress bar
    for batch_i, (im, targets, paths, shapes) in enumerate(pbar):
        callbacks.run("on_val_batch_start")
        with dt[0]:
            if cuda:
                im = im.to(device, non_blocking=True)
                targets = targets.to(device)
            im = im.half() if half else im.float()  # uint8 to fp16/32
            im /= 255  # 0 - 255 to 0.0 - 1.0
            nb, _, height, width = im.shape  # batch size, channels, height, width

        # Inference
        with dt[1]:
            if compute_loss:
                preds, train_out, mask = model(im)
            else:
                preds = model(im, augment=augment)
                train_out, mask = None, None
                if isinstance(preds, (list, tuple)) and len(preds) == 3:
                    preds, train_out, mask = preds

        # Loss
        if compute_loss:
            loss += compute_loss(train_out, mask, targets)[1]  # box, obj, cls

        # NMS
        targets[:, 2:] *= torch.tensor((width, height, width, height), device=device)  # to pixels
        lb = [targets[targets[:, 0] == i, 1:] for i in range(nb)] if save_hybrid else []  # for autolabelling
        with dt[2]:
            preds = non_max_suppression(
                preds, conf_thres, iou_thres, labels=lb, multi_label=True, agnostic=single_cls, max_det=max_det
            )

        # Metrics
        for si, pred in enumerate(preds):
            labels = targets[targets[:, 0] == si, 1:]
            nl, npr = labels.shape[0], pred.shape[0]  # number of labels, predictions
            path, shape = Path(paths[si]), shapes[si][0]
            correct = torch.zeros(npr, niou, dtype=torch.bool, device=device)  # init
            seen += 1

            if npr == 0:
                if task in ("val", "test") and not training:
                    add_race_stats(race_stats, race_match_stats(pred, labels, nc))
                if nl:
                    stats.append((correct, *torch.zeros((2, 0), device=device), labels[:, 0]))
                    if plots:
                        confusion_matrix.process_batch(detections=None, labels=labels[:, 0])
                continue

            # Predictions
            if single_cls:
                pred[:, 5] = 0
            predn = pred.clone()
            scale_boxes(im[si].shape[1:], predn[:, :4], shape, shapes[si][1])  # native-space pred

            # Evaluate
            if nl:
                tbox = xywh2xyxy(labels[:, 1:5])  # target boxes
                scale_boxes(im[si].shape[1:], tbox, shape, shapes[si][1])  # native-space labels
                labelsn = torch.cat((labels[:, 0:1], tbox), 1)  # native-space labels
                correct = process_batch(predn, labelsn, iouv)
                if task in ("val", "test") and not training:
                    add_race_stats(race_stats, race_match_stats(predn, labelsn, nc))
                if plots:
                    confusion_matrix.process_batch(predn, labelsn)
            elif task in ("val", "test") and not training:
                add_race_stats(race_stats, race_match_stats(predn, labels, nc))
            stats.append((correct, pred[:, 4], pred[:, 5], labels[:, 0]))  # (correct, conf, pcls, tcls)

            # Save/log
            if save_txt:
                (save_dir / "labels").mkdir(parents=True, exist_ok=True)
                save_one_txt(predn, save_conf, shape, file=save_dir / "labels" / f"{path.stem}.txt")
            if save_json:
                save_one_json(predn, jdict, path, class_map)  # append to COCO-JSON dictionary
            callbacks.run("on_val_image_end", pred, predn, path, names, im[si])

        # Plot images
        if plots:# and batch_i < 3:
            plot_images(im, targets, paths, save_dir / f"val_batch{batch_i}_labels.jpg", names)  # labels
            plot_images(im, output_to_target(preds), paths, save_dir / f"val_batch{batch_i}_pred.jpg", names)  # pred

        callbacks.run("on_val_batch_end", batch_i, im, targets, paths, shapes, preds)

    # Compute metrics
    stats = [torch.cat(x, 0).cpu().numpy() for x in zip(*stats)]  # to numpy
    if len(stats) and stats[0].any():
        tp, fp, p, r, f1, ap, ap_class = ap_per_class(*stats, plot=plots, save_dir=save_dir, names=names)
        ap50, ap = ap[:, 0], ap.mean(1)  # AP@0.5, AP@0.5:0.95
        mp, mr, map50, map = p.mean(), r.mean(), ap50.mean(), ap.mean()
    nt = np.bincount(stats[3].astype(int), minlength=nc)  # number of targets per class

    # Print results
    pf = "%22s" + "%11i" * 2 + "%11.3g" * 4  # print format
    LOGGER.info(pf % ("all", seen, nt.sum(), mp, mr, map50, map))
    if nt.sum() == 0:
        LOGGER.warning(f"WARNING ⚠️ no labels found in {task} set, can not compute metrics without labels")

    if task in ("val", "test") and not training:
        race_tp, race_fp, race_fn = (int(race_stats[k].sum()) for k in ("tp", "fp", "fn"))
        race_gt, race_recall, race_fdr = race_summary(race_tp, race_fp, race_fn)
        race_metrics = {
            "tp": race_tp,
            "fp": race_fp,
            "fn": race_fn,
            "gt": int(race_gt),
            "recall": race_recall,
            "fdr": race_fdr,
            "vehicle_iou": 0.35,
            "default_iou": 0.50,
            "vehicle_class": 24,
        }
        LOGGER.info(
            f"Race metrics: Recall={race_recall:.6f}, FDR={race_fdr:.6f}, "
            f"TP={race_tp}, FP={race_fp}, FN={race_fn}, GT={race_gt}"
        )

        group_defs = {"ship": range(0, 4), "aircraft": range(4, 24), "vehicle": range(24, 25)}
        group_rows = []
        LOGGER.info("Race metrics by group:")
        LOGGER.info(("%12s" + "%10s" * 7) % ("Group", "TP", "FP", "FN", "GT", "Recall", "FDR", "Pred"))
        for group, cls_ids in group_defs.items():
            cls_ids = list(cls_ids)
            tp_g = int(race_stats["tp"][cls_ids].sum())
            fp_g = int(race_stats["fp"][cls_ids].sum())
            fn_g = int(race_stats["fn"][cls_ids].sum())
            gt_g, recall_g, fdr_g = race_summary(tp_g, fp_g, fn_g)
            pred_g = tp_g + fp_g
            group_rows.append((group, tp_g, fp_g, fn_g, int(gt_g), recall_g, fdr_g, pred_g))
            LOGGER.info(
                ("%12s" + "%10i" * 4 + "%10.4f" * 2 + "%10i")
                % (group, tp_g, fp_g, fn_g, int(gt_g), recall_g, fdr_g, pred_g)
            )

        class_rows = []
        LOGGER.info("Race metrics by class:")
        LOGGER.info(
            ("%4s %20s" + "%8s" * 10)
            % ("ID", "Class", "TP", "FP", "FN", "GT", "Recall", "FDR", "Wrong", "Dup", "Bg/IoU", "Pred")
        )
        for c in range(nc):
            tp_c, fp_c, fn_c = (int(race_stats[k][c]) for k in ("tp", "fp", "fn"))
            gt_c, recall_c, fdr_c = race_summary(tp_c, fp_c, fn_c)
            wrong_c = int(race_stats["fp_wrong_class"][c])
            dup_c = int(race_stats["fp_duplicate"][c])
            bg_iou_c = int(race_stats["fp_background_or_iou"][c])
            pred_c = tp_c + fp_c
            class_name = names.get(c, str(c))
            class_rows.append((c, class_name, tp_c, fp_c, fn_c, int(gt_c), recall_c, fdr_c, wrong_c, dup_c, bg_iou_c, pred_c))
            LOGGER.info(
                ("%4i %20s" + "%8i" * 4 + "%8.4f" * 2 + "%8i" * 4)
                % (c, class_name, tp_c, fp_c, fn_c, int(gt_c), recall_c, fdr_c, wrong_c, dup_c, bg_iou_c, pred_c)
            )

        with open(save_dir / "race_metrics.json", "w") as f:
            json.dump(race_metrics, f, indent=2)
        with open(save_dir / "race_metrics.csv", "w") as f:
            f.write("tp,fp,fn,gt,recall,fdr,vehicle_iou,default_iou,vehicle_class\n")
            f.write(
                f"{race_tp},{race_fp},{race_fn},{race_gt},{race_recall:.10f},{race_fdr:.10f},"
                f"0.35,0.50,24\n"
            )
        with open(save_dir / "race_metrics_by_group.csv", "w") as f:
            f.write("group,tp,fp,fn,gt,recall,fdr,pred\n")
            for row in group_rows:
                f.write("%s,%i,%i,%i,%i,%.10f,%.10f,%i\n" % row)
        with open(save_dir / "race_metrics_by_class.csv", "w") as f:
            f.write("class_id,class_name,tp,fp,fn,gt,recall,fdr,fp_wrong_class,fp_duplicate,fp_background_or_iou,pred\n")
            for row in class_rows:
                f.write("%i,%s,%i,%i,%i,%i,%.10f,%.10f,%i,%i,%i,%i\n" % row)
        with open(save_dir / "race_confusion_matrix.csv", "w") as f:
            f.write("true_class_id,true_class_name,pred_class_id,pred_class_name,count\n")
            confusion = race_stats["confusion"]
            for true_c, pred_c in zip(*np.nonzero(confusion)):
                f.write(
                    f"{true_c},{names.get(int(true_c), str(true_c))},{pred_c},{names.get(int(pred_c), str(pred_c))},{int(confusion[true_c, pred_c])}\n"
                )

    # Print results per class
    # if (verbose or (nc < 50 and not training)) and nc > 1 and len(stats):
    #     for i, c in enumerate(ap_class):
    #         LOGGER.info(pf % (names[c], seen, nt[c], p[i], r[i], ap50[i], ap[i]))

    # Print speeds
    t = tuple(x.t / seen * 1e3 for x in dt)  # speeds per image
    if not training:
        shape = (batch_size, 3, imgsz, imgsz)
        LOGGER.info(f"Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS per image at shape {shape}" % t)

    ######################################## out put ####################################################################
    import openpyxl
    def save_to_excel(data, file_name='output.xlsx'):
        # 尝试加载现有的 Excel 文件，如果不存在则创建一个新的
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
        
        # 找到下一行空行
        next_row = sheet.max_row + 1
        
        # 将数据写入下一行
        if isinstance(data, list):
            sheet.append(data)
        else:
            sheet.append([data])
        
        # 保存 Excel 文件
        workbook.save(file_name)

    output_excel = save_dir / 'result.xlsx'
    s = ('%20s' + '%12s' * 7) % ('Class', 'P', 'R', 'f1', 'mAP@0.5', 'mAP@0.75', 'mAP@.90', 'mAP@.5:.95')
    print(s)
    tp, fp, p, r, f1, ap, ap_class = ap_per_class(*stats, plot=plots, save_dir=save_dir, names=names)
    ap50, ap75, ap90, ap = ap[:, 0], ap[:, 5], ap[:, 9], ap.mean(1)  # AP@0.5, AP@0.5:0.95
    save_to_excel(['Class', 'P', 'R', 'f1', 'mAP@0.5', 'mAP@0.75', 'mAP@.90', 'mAP@.5:.95'], output_excel)
    mp, mr, mf1, map50, map75, map95, map = p.mean(), r.mean(), f1.mean(), ap50.mean(), ap75.mean(), ap90.mean(), ap.mean()

    # Print results
    pf = '%20s' + '%12.3g' * 7  # print format
    print(pf % ('all', mp, mr, mf1, map50, map75, map95, map))
    save_to_excel(['all', mp, mr, mf1, map50, map75, map95, map], output_excel)

    # Print results per class
    for i, c in enumerate(ap_class):
        print(pf % (names[i], p[i], r[i], f1[i], ap50[i], ap75[i], ap90[i], ap[i]))
        save_to_excel([names[i], p[i], r[i], f1[i], ap50[i], ap75[i], ap90[i], ap[i]], output_excel)
    ######################################## End out put ####################################################################

    # Plots
    if plots:
        confusion_matrix.plot(save_dir=save_dir, names=list(names.values()))
        callbacks.run("on_val_end", nt, tp, fp, p, r, f1, ap, ap50, ap_class, confusion_matrix)

    # Save JSON
    if save_json and len(jdict):
        w = Path(weights[0] if isinstance(weights, list) else weights).stem if weights is not None else ""  # weights
        anno_json = str(Path("../datasets/coco/annotations/instances_val2017.json"))  # annotations
        if not os.path.exists(anno_json):
            anno_json = os.path.join(data["path"], "annotations", "instances_val2017.json")
        pred_json = str(save_dir / f"{w}_predictions.json")  # predictions
        LOGGER.info(f"\nEvaluating pycocotools mAP... saving {pred_json}...")
        with open(pred_json, "w") as f:
            json.dump(jdict, f)

        try:  # https://github.com/cocodataset/cocoapi/blob/master/PythonAPI/pycocoEvalDemo.ipynb
            check_requirements("pycocotools>=2.0.6")
            from pycocotools.coco import COCO
            from pycocotools.cocoeval import COCOeval

            anno = COCO(anno_json)  # init annotations api
            pred = anno.loadRes(pred_json)  # init predictions api
            eval = COCOeval(anno, pred, "bbox")
            if is_coco:
                eval.params.imgIds = [int(Path(x).stem) for x in dataloader.dataset.im_files]  # image IDs to evaluate
            eval.evaluate()
            eval.accumulate()
            eval.summarize()
            map, map50 = eval.stats[:2]  # update results (mAP@0.5:0.95, mAP@0.5)
        except Exception as e:
            LOGGER.info(f"pycocotools unable to run: {e}")

    # Return results
    model.float()  # for training
    if not training:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ""
        LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
    maps = np.zeros(nc) + map
    for i, c in enumerate(ap_class):
        maps[c] = ap[i]
    return (mp, mr, map50, map, *(loss.cpu() / len(dataloader)).tolist()), maps, t


def parse_opt():
    """
    Parse command-line options for configuring YOLOv5 model inference.

    Args:
        data (str, optional): Path to the dataset YAML file. Default is 'data/coco128.yaml'.
        weights (list[str], optional): List of paths to model weight files. Default is 'yolov5s.pt'.
        batch_size (int, optional): Batch size for inference. Default is 32.
        imgsz (int, optional): Inference image size in pixels. Default is 640.
        conf_thres (float, optional): Confidence threshold for predictions. Default is 0.001.
        iou_thres (float, optional): IoU threshold for Non-Max Suppression (NMS). Default is 0.6.
        max_det (int, optional): Maximum number of detections per image. Default is 300.
        task (str, optional): Task type - options are 'train', 'val', 'test', 'speed', or 'study'. Default is 'val'.
        device (str, optional): Device to run the model on. e.g., '0' or '0,1,2,3' or 'cpu'. Default is empty to let the system choose automatically.
        workers (int, optional): Maximum number of dataloader workers per rank in DDP mode. Default is 8.
        single_cls (bool, optional): If set, treats the dataset as a single-class dataset. Default is False.
        augment (bool, optional): If set, performs augmented inference. Default is False.
        verbose (bool, optional): If set, reports mAP by class. Default is False.
        save_txt (bool, optional): If set, saves results to *.txt files. Default is False.
        save_hybrid (bool, optional): If set, saves label+prediction hybrid results to *.txt files. Default is False.
        save_conf (bool, optional): If set, saves confidences in --save-txt labels. Default is False.
        save_json (bool, optional): If set, saves results to a COCO-JSON file. Default is False.
        project (str, optional): Project directory to save results to. Default is 'runs/val'.
        name (str, optional): Name of the directory to save results to. Default is 'exp'.
        exist_ok (bool, optional): If set, existing directory will not be incremented. Default is False.
        half (bool, optional): If set, uses FP16 half-precision inference. Default is False.
        dnn (bool, optional): If set, uses OpenCV DNN for ONNX inference. Default is False.

    Returns:
        argparse.Namespace: Parsed command-line options.

    Notes:
        - The '--data' parameter is checked to ensure it ends with 'coco.yaml' if '--save-json' is set.
        - The '--save-txt' option is set to True if '--save-hybrid' is enabled.
        - Args are printed using `print_args` to facilitate debugging.

    Example:
        To validate a trained YOLOv5 model on a COCO dataset:
        ```python
        $ python val.py --weights yolov5s.pt --data coco128.yaml --img 640
        ```
        Different model formats could be used instead of `yolov5s.pt`:
        ```python
        $ python val.py --weights yolov5s.pt yolov5s.torchscript yolov5s.onnx yolov5s_openvino_model yolov5s.engine
        ```
        Additional options include saving results in different formats, selecting devices, and more.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", type=str, default=ROOT / "data/competition25.yaml", help="dataset.yaml path")
    parser.add_argument("--name", default="competition25", help="save to project/name")

    parser.add_argument("--weights", nargs="+", type=str, default=ROOT / "./checkpoints/best.pt", help="model path(s)")
    parser.add_argument("--batch-size", type=int, default=1, help="batch size")
    parser.add_argument("--imgsz", "--img", "--img-size", type=int, default=512, help="inference size (pixels)")
    parser.add_argument("--conf-thres", type=float, default=0.001, help="confidence threshold")
    parser.add_argument("--iou-thres", type=float, default=0.6, help="NMS IoU threshold")
    parser.add_argument("--max-det", type=int, default=300, help="maximum detections per image")
    parser.add_argument("--task", default="val", help="train, val, test, speed or study")
    parser.add_argument("--device", default="0", help="cuda device, i.e. 0 or 0,1,2,3 or cpu")
    parser.add_argument("--workers", type=int, default=2, help="max dataloader workers (per RANK in DDP mode)")
    parser.add_argument("--single-cls", action="store_true", help="treat as single-class dataset")
    parser.add_argument("--augment", action="store_true", help="augmented inference")
    parser.add_argument("--verbose", action="store_true", help="report mAP by class")
    parser.add_argument("--save-txt", action="store_true", help="save results to *.txt")
    parser.add_argument("--save-hybrid", action="store_true", help="save label+prediction hybrid results to *.txt")
    parser.add_argument("--save-conf", action="store_true", help="save confidences in --save-txt labels")
    parser.add_argument("--save-json", action="store_true", help="save a COCO-JSON results file")
    parser.add_argument("--project", default=ROOT / "runs/val", help="save to project/name")
    parser.add_argument("--exist-ok", action="store_true", help="existing project/name ok, do not increment")
    parser.add_argument("--half", action="store_true", help="use FP16 half-precision inference")
    parser.add_argument("--dnn", action="store_true", help="use OpenCV DNN for ONNX inference")
    opt = parser.parse_args()
    opt.data = check_yaml(opt.data)  # check YAML
    opt.save_json |= opt.data.endswith("coco.yaml")
    opt.save_txt |= opt.save_hybrid
    print_args(vars(opt))
    return opt


def main(opt):
    """
    Executes YOLOv5 tasks like training, validation, testing, speed, and study benchmarks based on provided options.

    Args:
        opt (argparse.Namespace): Parsed command-line options.
            This includes values for parameters like 'data', 'weights', 'batch_size', 'imgsz', 'conf_thres',
            'iou_thres', 'max_det', 'task', 'device', 'workers', 'single_cls', 'augment', 'verbose', 'save_txt',
            'save_hybrid', 'save_conf', 'save_json', 'project', 'name', 'exist_ok', 'half', and 'dnn', essential
            for configuring the YOLOv5 tasks.

    Returns:
        None

    Examples:
        To validate a trained YOLOv5 model on the COCO dataset with a specific weights file, use:
        ```python
        $ python val.py --weights yolov5s.pt --data coco128.yaml --img 640
        ```
    """
    check_requirements(ROOT / "requirements.txt", exclude=("tensorboard", "thop"))

    if opt.task in ("train", "val", "test"):  # run normally
        if opt.conf_thres > 0.001:  # https://github.com/ultralytics/yolov5/issues/1466
            LOGGER.info(f"WARNING ⚠️ confidence threshold {opt.conf_thres} > 0.001 produces invalid results")
        if opt.save_hybrid:
            LOGGER.info("WARNING ⚠️ --save-hybrid will return high mAP from hybrid labels, not from predictions alone")
        run(**vars(opt))

    else:
        weights = opt.weights if isinstance(opt.weights, list) else [opt.weights]
        opt.half = torch.cuda.is_available() and opt.device != "cpu"  # FP16 for fastest results
        if opt.task == "speed":  # speed benchmarks
            # python val.py --task speed --data coco.yaml --batch 1 --weights yolov5n.pt yolov5s.pt...
            opt.conf_thres, opt.iou_thres, opt.save_json = 0.25, 0.45, False
            for opt.weights in weights:
                run(**vars(opt), plots=False)

        elif opt.task == "study":  # speed vs mAP benchmarks
            # python val.py --task study --data coco.yaml --iou 0.7 --weights yolov5n.pt yolov5s.pt...
            for opt.weights in weights:
                f = f"study_{Path(opt.data).stem}_{Path(opt.weights).stem}.txt"  # filename to save to
                x, y = list(range(256, 1536 + 128, 128)), []  # x axis (image sizes), y axis
                for opt.imgsz in x:  # img-size
                    LOGGER.info(f"\nRunning {f} --imgsz {opt.imgsz}...")
                    r, _, t = run(**vars(opt), plots=False)
                    y.append(r + t)  # results and times
                np.savetxt(f, y, fmt="%10.4g")  # save
            subprocess.run(["zip", "-r", "study.zip", "study_*.txt"])
            plot_val_study(x=x)  # plot
        else:
            raise NotImplementedError(f'--task {opt.task} not in ("train", "val", "test", "speed", "study")')


if __name__ == "__main__":
    opt = parse_opt()
    main(opt)
